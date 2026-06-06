"""Convert the OTRF Mordor APT3 emulation playbook (xlsx) into a committed JSON
of authoritative attack steps — the SOURCE of ground-truth labels.

The playbook (data/ground_truth/apt3_mordor_playbook.xlsx, pulled from
OTRF/Security-Datasets) documents every operator step of the MITRE APT3 Round-1
evaluation: the command run, its ATT&CK technique, the host and user. We extract
the process-spawning `shell <cmd>` steps so the labeler can match them against
log events. Run once per dataset (re-run if the playbook changes).

This replaces hand-written pattern heuristics, which under-labeled real attack
steps (e.g. `net localgroup Administrators` — a documented APT3 discovery step —
was labeled benign because it lacked `/domain`).
"""
import json
import re
from pathlib import Path

import openpyxl

GT_DIR = Path(__file__).parent.parent / "data" / "ground_truth"


def build(xlsx_name: str = "apt3_mordor_playbook.xlsx",
          out_name: str = "apt3_attack_steps.json") -> int:
    wb = openpyxl.load_workbook(GT_DIR / xlsx_name, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    ci = {h: i for i, h in enumerate(rows[0])}

    commands, artifacts = [], []
    seen_cmd, seen_art = set(), set()
    for r in rows[1:]:
        cmds = str(r[ci["Empire Commands / Notes"]] or "")
        tech = r[ci["Technique Id"]]
        meta = {"step": r[ci["ATT&CK Eval Step"]], "technique_id": tech,
                "technique_name": r[ci["Technique Name"]], "tactic": r[ci["Tactic"]],
                "source_username": r[ci["Source Username"]]}
        for line in cmds.split("\n"):
            line = line.strip()
            # (1) process-spawning OS commands
            m = re.match(r"shell\s+(.+)", line, re.I)
            if m and (m.group(1).strip(), tech) not in seen_cmd:
                seen_cmd.add((m.group(1).strip(), tech))
                commands.append({"command": m.group(1).strip(), **meta})
            # (2) payload artifacts dropped by stagers / uploads (then executed as
            #     wscript/cmd/powershell processes referencing the file)
            for am in re.finditer(r"(?:OutFile|upload)\s+(\S+)", line, re.I):
                fname = am.group(1).rsplit("/", 1)[-1].rsplit("\\", 1)[-1].lower()
                if "." in fname and (fname, tech) not in seen_art:
                    seen_art.add((fname, tech))
                    artifacts.append({"artifact": fname, **meta})

    out_path = GT_DIR / out_name
    payload = {"commands": commands, "artifacts": artifacts}
    out_path.write_text(json.dumps(payload, indent=2))
    print(f"wrote {len(commands)} shell commands + {len(artifacts)} payload artifacts -> {out_path}")
    return len(commands) + len(artifacts)


if __name__ == "__main__":
    build()
